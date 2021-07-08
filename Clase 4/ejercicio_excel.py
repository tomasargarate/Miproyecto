from openpyxl import load_workbook

wb = load_workbook('lista_de_precios.xlsx', data_only=True)
sheets = wb.sheetnames
valores_precio = []
for hoja in sheets:
   hoja_activa = wb[hoja]
   valor = hoja_activa['F10'].value
   valores_precio.append(round(valor,2))
